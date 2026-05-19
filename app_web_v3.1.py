import streamlit as st
import streamlit.components.v1 as components # NOVO: Importação para injetar o bloqueio do teclado
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook, load_workbook
import pandas as pd 

# =============================================================
# CARREGANDO DADOS REGISTRADOS NO BANCO DE DADOS
# =============================================================

caminh_arquivo = 'base_dados_partido.xlsx'

# 2. Importamos cada tabela (aba) para um DataFrame diferente
# Usamos engine='openpyxl' para garantir a leitura correta do formato .xlsx [10]
df_receitas = pd.read_excel(caminh_arquivo, sheet_name='receitas', engine='openpyxl')
df_despesas = pd.read_excel(caminh_arquivo, sheet_name='despesas', engine='openpyxl')
df_usuarios = pd.read_excel(caminh_arquivo, sheet_name='usuarios', engine='openpyxl')
df_bancos = pd.read_excel(caminh_arquivo, sheet_name='bancos', engine='openpyxl')
df_credores = pd.read_excel(caminh_arquivo, sheet_name='credores', engine='openpyxl')
df_colaboradores = pd.read_excel(caminh_arquivo, sheet_name='colaboradores', engine='openpyxl')

# =============================================================
# CARREGANDO DADOS DA PRESTAÇÃO DE CONTAS DE 2024
# =============================================================
extrato_bancario_pca = "extrato_partido_.xlsx"
extrato_bancario_pca = pd.read_excel(extrato_bancario_pca)

# =============================================================
# CONFIGURAÇÃO INICIAL DA PÁGINA WEB
# =============================================================
st.set_page_config(page_title="Gestão Partidária", page_icon="🏛️", layout="centered")

# =============================================================
# FUNÇÃO PARA BLOQUEAR O "ENTER" NOS FORMULÁRIOS
# =============================================================
def bloquear_enter():
    """Injeta um script no navegador para impedir o envio de forms com a tecla Enter"""
    components.html(
        """
        <script>
        const doc = window.parent.document;
        doc.addEventListener('keydown', function(e) {
            // Se a tecla for Enter e o usuário estiver digitando em um campo (INPUT)
            if (e.key === 'Enter' && e.target.nodeName === 'INPUT') {
                e.preventDefault(); // Cancela o envio
            }
        });
        </script>
        """,
        height=0, width=0 # Mantém o script invisível na tela
    )

# =============================================================
# CRIANDO ESTRUTURA DO BANCO EM EXCEL (Lógica Intacta)
# =============================================================
ARQUIVO_BD = Path("base_dados_partido.xlsx")

TABELAS = {
    "receitas": ["id", "data_registro_sistema", "data_receita", "agencia", "conta_corrente", "origem_receita", "agencia_origem", "conta_origem", "natureza_receita", "valor", "nota_explicativa"],
    "despesas": ["id", "data_registro_sistema", "data_despesa", "processo", "contrato", "cnpj_cpf", "agencia_pagadora", "conta_pagadora", "natureza_despesa", "valor", "nota_explicativa"],
    "usuarios": ["id", "data_registro_sistema", "nome_completo", "data_nascimento", "cpf", "estado", "cidade", "rua", "numero", "bairro", "cep"],
    "bancos": ["id", "data_registro_sistema", "nome_banco", "numero_agencia", "numero_conta_corrente", "nome_conta"],
    "credores": ["id", "data_registro_sistema", "cnpj_cpf", "nome_credor", "estado", "cidade", "rua", "numero", "cep", "agencia_credor", "conta_credor"],
    "colaboradores": ["id", "data_registro_sistema", "cpf", "data_nascimento", "nome_completo", "estado", "cidade", "rua", "numero", "cep", "carteira_trabalho"]
}

def inicializar_arquivo_excel():
    if not ARQUIVO_BD.exists():
        wb = Workbook()
        wb.remove(wb.active)
        for nome_aba, cabecalhos in TABELAS.items():
            ws = wb.create_sheet(nome_aba)
            ws.append(cabecalhos)
        wb.save(ARQUIVO_BD)
        return

    wb = load_workbook(ARQUIVO_BD)
    precisa_salvar = False
    for nome_aba, cabecalhos in TABELAS.items():
        if nome_aba not in wb.sheetnames:
            ws = wb.create_sheet(nome_aba)
            ws.append(cabecalhos)
            precisa_salvar = True
    if precisa_salvar:
        wb.save(ARQUIVO_BD)

def proximo_id(ws):
    if ws.max_row == 1: return 1
    ultimo_id = ws.cell(row=ws.max_row, column=1).value
    try:
        return int(ultimo_id) + 1
    except:
        return ws.max_row

def salvar_registro_excel(nome_aba, dados):
    inicializar_arquivo_excel()
    wb = load_workbook(ARQUIVO_BD)
    ws = wb[nome_aba]
    novo_id = proximo_id(ws)

    dados_completos = {
        "id": novo_id,
        "data_registro_sistema": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        **dados
    }
    
    linha = [dados_completos.get(coluna, "") for coluna in TABELAS[nome_aba]]
    ws.append(linha)
    wb.save(ARQUIVO_BD)

def carregar_contas_bancarias():
    inicializar_arquivo_excel()
    wb = load_workbook(ARQUIVO_BD)
    if "bancos" not in wb.sheetnames: return []
    ws = wb["bancos"]
    if ws.max_row <= 1: return []
    
    cabecalhos = [str(c.value).strip() if c.value else "" for c in ws[1]]
    return [dict(zip(cabecalhos, linha)) for linha in ws.iter_rows(min_row=2, values_only=True) if dict(zip(cabecalhos, linha)).get("numero_agencia")]

def carregar_credores():
    inicializar_arquivo_excel()
    wb = load_workbook(ARQUIVO_BD)
    if "credores" not in wb.sheetnames: return []
    ws = wb["credores"]
    if ws.max_row <= 1: return []
    
    cabecalhos = [str(c.value).strip() if c.value else "" for c in ws[1]]
    return [dict(zip(cabecalhos, linha)) for linha in ws.iter_rows(min_row=2, values_only=True) if dict(zip(cabecalhos, linha)).get("cnpj_cpf")]


# =============================================================
# INTERFACES WEB (AS TELAS DO STREAMLIT)
# =============================================================

def tela_receita():
    st.header("Registrar Receita")
    
    contas_cadastradas = carregar_contas_bancarias()
    if not contas_cadastradas:
        st.warning("Não existe conta bancária cadastrada. Cadastre uma conta no menu 'Cadastrar Banco' primeiro.")
        return

    opcoes_contas = [f"Agência: {c['numero_agencia']} | Conta: {c['numero_conta_corrente']} | {c['nome_banco']} - {c['nome_conta']}" for c in contas_cadastradas]
    
    conta_selecionada = st.selectbox("Selecione a Conta Creditada:", opcoes_contas)
    
    with st.form("form_receita", clear_on_submit=True):
        col1, col2 = st.columns(2) 
        
        with col1:
            st.subheader("Dados da Receita")
            data_rec = st.date_input("Data da Receita", format="DD/MM/YYYY")
            natureza_rec = st.text_input("Tipo (natureza) da Receita")
            valor_rec = st.number_input("Valor (R$)", min_value=0.0, format="%.2f")
            
        with col2:
            st.subheader("Fonte Pagadora")
            origem_rec = st.text_input("Origem da Receita")
            agencia_origem = st.text_input("Agência de Origem")
            conta_origem = st.text_input("Conta Corrente de Origem")
            
        nota_explicativa = st.text_area("Informações Adicionais", height=100)
        
        submit = st.form_submit_button("Salvar Receita", type="primary")
        
        if submit:
            partes_conta = conta_selecionada.split(" | ")
            agencia = partes_conta[0].replace("Agência: ", "")
            conta = partes_conta[1].replace("Conta: ", "")
            
            dados = {
                "data_receita": data_rec.strftime("%d/%m/%Y"),
                "agencia": agencia,
                "conta_corrente": conta,
                "origem_receita": origem_rec,
                "agencia_origem": agencia_origem,
                "conta_origem": conta_origem,
                "natureza_receita": natureza_rec,
                "valor": f"R$ {valor_rec:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."),
                "nota_explicativa": nota_explicativa
            }
            salvar_registro_excel("receitas", dados)
            st.success("Receita registrada com sucesso!")


def tela_despesa():
    st.header("Registrar Despesa")
    
    # 1. CARREGAMENTO E VALIDAÇÃO DE DADOS BÁSICOS
    credores_cadastrados = carregar_credores()
    if not credores_cadastrados: # Verifica se há credores cadastrados, caso contrário, exibe um aviso de usuário não cadastrado.
        st.warning("Não existe credor cadastrado no sistema. Cadastre um no menu lateral.") 
        return

    contas_cadastradas = carregar_contas_bancarias()
    if not contas_cadastradas:
        st.warning("Não existe conta bancária cadastrada. Cadastre uma conta no menu 'Cadastrar Banco' primeiro.")
        return

    # 2. PREPARAÇÃO DAS LISTAS DE OPÇÕES
    opcoes_credores = [f"{c['cnpj_cpf']} - {c['nome_credor']}" for c in credores_cadastrados]
    opcoes_contas = [f"Agência: {c['numero_agencia']} | Conta: {c['numero_conta_corrente']} | {c['nome_banco']} - {c['nome_conta']}" for c in contas_cadastradas]

    # Seleção do Credor
    # Note que o selectbox de seleção de credor é construído ANTES do formulário.
    credor_selecionado = st.selectbox("Selecione o Credor (CNPJ/CPF):", opcoes_credores)
    
    # 3. CONSTRUÇÃO DO FORMULÁRIO
    with st.form("form_despesa", clear_on_submit=True):
        col1, col2 = st.columns(2)
        
        with col1:
            data_desp = st.date_input("Data da Despesa", format="DD/MM/YYYY")
            processo = st.text_input("Processo nº")
            contrato = st.text_input("Contrato nº")
            
        with col2:
            natureza_desp = st.text_input("Tipo (natureza) da Despesa")
            valor_desp = st.number_input("Valor (R$)", min_value=0.0, format="%.2f")
            
        st.divider() # Adiciona uma linha horizontal para separar os assuntos visualmente
        st.subheader("Pagamento")
        
        # A CAIXA DE SELEÇÃO DO BANCO ENTRA AQUI
        conta_selecionada = st.selectbox("Selecione a Conta Debitada (Saída do Dinheiro):", opcoes_contas)

        # Caixa de texto para Justificativa, que pode ser usada para detalhar a despesa, justificar o motivo, ou qualquer informação adicional relevante.
        nota_explicativa = st.text_area("Justificativa")
        
        # O BOTÃO DE ENVIO DO FORMULÁRIO FICA POR ÚLTIMO, APÓS TODAS AS SELEÇÕES E CAMPOS DE TEXTO
        submit = st.form_submit_button("Salvar Despesa", type="primary")
        
        # 4. SALVAMENTO UNIFICADO
        if submit:
            # Extrai apenas o CNPJ
            cnpj_puro = credor_selecionado.split(" - ")[0]
            
            # Extrai apenas a Agência e a Conta da opção selecionada
            partes_conta = conta_selecionada.split(" | ")
            agencia_deb = partes_conta[0].replace("Agência: ", "")
            conta_deb = partes_conta[1].replace("Conta: ", "")
            
            # Reúne TODOS os dados em um único dicionário para a tabela
            dados = {
                "data_despesa": data_desp.strftime("%d/%m/%Y"),
                "processo": processo,
                "contrato": contrato,
                "cnpj_cpf": cnpj_puro,
                "agencia_pagadora": agencia_deb,
                "conta_pagadora": conta_deb,
                "natureza_despesa": natureza_desp,
                "valor": f"R$ {valor_desp:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."),
                "nota_explicativa": nota_explicativa
            }
            
            # Salva na aba 'despesas'
            salvar_registro_excel("despesas", dados)
            st.success("Despesa registrada com sucesso!")


def tela_cadastro_usuario():
    st.header("Cadastrar Usuário")
    with st.form("form_usuario", clear_on_submit=True):
        nome = st.text_input("Nome Completo")
        col1, col2 = st.columns(2)
        with col1:
            data_nasc = st.date_input("Data de Nascimento", format="DD/MM/YYYY")
            cpf = st.text_input("CPF (Apenas números)", max_chars=11)
            cep = st.text_input("CEP (Apenas números)", max_chars=8)
        with col2:
            estado = st.text_input("Estado (UF)", max_chars=2)
            cidade = st.text_input("Cidade")
            bairro = st.text_input("Bairro")
            
        col3, col4 = st.columns([3, 1]) 
        with col3:
            rua = st.text_input("Rua")
        with col4:
            numero = st.text_input("Número")
            
        if st.form_submit_button("Cadastrar Usuário", type="primary"):
            dados = {
                "nome_completo": nome, "data_nascimento": data_nasc.strftime("%d/%m/%Y"),
                "cpf": cpf, "estado": estado, "cidade": cidade,
                "rua": rua, "numero": numero, "bairro": bairro, "cep": cep
            }
            salvar_registro_excel("usuarios", dados)
            st.success("Usuário cadastrado com sucesso!")
    
    st.header("Usuários Cadastrados")
    st.dataframe(df_usuarios)


def tela_cadastro_banco():
    st.header("Cadastrar Dados Bancários")
    with st.form("form_banco", clear_on_submit=True):
        nome_banco = st.text_input("Nome do Banco")
        agencia = st.text_input("Número da Agência")
        conta = st.text_input("Nº Conta Corrente")
        nome_conta = st.text_input("Nome da Conta (Ex: Doações, Fundo)")
        
        if st.form_submit_button("Cadastrar Banco", type="primary"):
            salvar_registro_excel("bancos", {"nome_banco": nome_banco, "numero_agencia": agencia, "numero_conta_corrente": conta, "nome_conta": nome_conta})
            st.success("Banco cadastrado com sucesso!")
    
    st.header("Bancos Cadastrados")
    st.dataframe(df_bancos)


def tela_cadastro_credor():
    st.header("Cadastrar Credor")
    with st.form("form_credor", clear_on_submit=True):
        col1, col2 = st.columns(2)
        with col1:
            cnpj = st.text_input("CNPJ/CPF (Apenas números)")
            nome = st.text_input("Nome do Credor")
            agencia = st.text_input("Agência (Credor)")
            conta = st.text_input("Conta Corrente (Credor)")
        with col2:
            estado = st.text_input("Estado (UF)")
            cidade = st.text_input("Cidade")
            rua = st.text_input("Rua")
            numero = st.text_input("Número")
            cep = st.text_input("CEP")
            
        if st.form_submit_button("Cadastrar Credor", type="primary"):
            dados = {"cnpj_cpf": cnpj, "nome_credor": nome, "estado": estado, "cidade": cidade, "rua": rua, "numero": numero, "cep": cep, "agencia_credor": agencia, "conta_credor": conta}
            salvar_registro_excel("credores", dados)
            st.success("Credor cadastrado com sucesso!")

    st.header("Credores Cadastrados")
    st.dataframe(df_credores)

def tela_cadastro_colaborador():
    st.header("Cadastrar Colaborador")
    with st.form("form_colab", clear_on_submit=True):
        nome = st.text_input("Nome Completo")
        col1, col2 = st.columns(2)
        with col1:
            cpf = st.text_input("CPF")
            data_nasc = st.date_input("Data de Nascimento", format="DD/MM/YYYY")
            ctb = st.text_input("Carteira de Trabalho")
            cep = st.text_input("CEP")
        with col2:
            estado = st.text_input("Estado (UF)")
            cidade = st.text_input("Cidade")
            rua = st.text_input("Rua")
            numero = st.text_input("Número")
            
        if st.form_submit_button("Cadastrar Colaborador", type="primary"):
            dados = {"cpf": cpf, "data_nascimento": data_nasc.strftime("%d/%m/%Y"), "nome_completo": nome, "estado": estado, "cidade": cidade, "rua": rua, "numero": numero, "cep": cep, "carteira_trabalho": ctb}
            salvar_registro_excel("colaboradores", dados)
            st.success("Colaborador cadastrado com sucesso!")


# =============================================================
# O MOTOR DO PROGRAMA (MENU LATERAL)
# =============================================================
def main():
    bloquear_enter() # Chama a função que trava o Enter nos formulários do site
    inicializar_arquivo_excel() # Garante que o banco existe
    
    st.sidebar.title("PSD")
    st.sidebar.markdown("Contábil - (Receita x Despesa)")
    
    menu = st.sidebar.radio(
        "Selecione uma opção:",
        ["Registrar Receita", "Registrar Despesa", "Cadastrar Usuário", "Cadastrar Banco", "Cadastrar Credor", "Cadastro de Colaboradores"]
    )
    
    if menu == "Registrar Receita":
        tela_receita()
    elif menu == "Registrar Despesa":
        tela_despesa()
    elif menu == "Cadastrar Usuário":
        tela_cadastro_usuario()
    elif menu == "Cadastrar Banco":
        tela_cadastro_banco()
    elif menu == "Cadastrar Credor":
        tela_cadastro_credor()
    elif menu == "Cadastro de Colaboradores":
        tela_cadastro_colaborador()

# Comando para iniciar o Streamlit
if __name__ == "__main__":
    main()
